"""
Generador de Plan de Muestreo — interfaz web local.
Ejecutar: python app.py  → abrir http://localhost:5000
"""

import json
import os
import tempfile
from pathlib import Path

import openpyxl
from flask import Flask, jsonify, render_template_string, request, send_file

from extraer_cotizacion import extraer_cotizacion
from generar_plan import generar_plan, nro_base

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

BASE = Path(__file__).resolve().parent
OUTPUT_DIR = BASE / "_planes_generados"
OUTPUT_DIR.mkdir(exist_ok=True)
MAPEO_PATH = BASE / "mapeo_parametros.json"
PLANTILLA   = BASE / "COT2026-0103 - SOCOSANI S A.xlsm"


def matrices_disponibles() -> list[str]:
    """Lee las matrices de la hoja MATRIZ del Excel, filtrando entradas que no son matrices."""
    wb = openpyxl.load_workbook(str(PLANTILLA), keep_vba=True, data_only=True)
    ws = wb["MATRIZ"]
    excluir = {"MATRICES", "Recuento de microorganismos aerobios mesófilos",
                "Recuento de mohos y levaduras",
                "ENUMERACIÓN DE BACTERIAS ANAEROBIAS SULFITO REDUCTORES"}
    resultado = []
    for row in ws.iter_rows(min_row=2, max_row=50):
        val = row[1].value  # columna B
        if val and str(val).strip() and str(val).strip() not in excluir:
            resultado.append(str(val).strip())
    return resultado


def params_info() -> dict:
    """Devuelve {grupo: {envase, volumen, preservacion}} desde la hoja PARAMETROS."""
    wb = openpyxl.load_workbook(str(PLANTILLA), keep_vba=True, data_only=True)
    ws = wb["PARAMETROS"]
    info = {}
    for row in ws.iter_rows(min_row=2, max_row=400):
        nombre = row[0].value
        if nombre and str(nombre).strip():
            clave = str(nombre).strip()
            info[clave] = {
                "envase":       str(row[1].value or "").strip(),   # B
                "volumen":      str(row[2].value or "").strip(),   # C
                "preservacion": str(row[3].value or "").strip(),   # D
            }
    return info


def grupos_disponibles() -> list[str]:
    wb = openpyxl.load_workbook(str(PLANTILLA), keep_vba=True, data_only=False)
    ws = wb["PARAMETROS"]
    vistos, resultado = set(), []
    for row in ws.iter_rows(min_row=2, max_row=400):
        v = row[0].value
        if v and str(v).strip() and str(v).strip() not in vistos:
            vistos.add(str(v).strip())
            resultado.append(str(v).strip())
    return sorted(resultado)


# ── HTML ─────────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <title>Generador de Plan de Muestreo</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:'Segoe UI',sans-serif;background:#f0f2f5;min-height:100vh;padding:32px 16px;display:flex;justify-content:center}
    .wrap{width:100%;max-width:800px}

    /* Card */
    .card{background:#fff;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.1);margin-bottom:20px;overflow:hidden}
    .card-head{background:#1a5276;color:#fff;padding:18px 24px}
    .card-head h1{font-size:1.15rem;font-weight:600}
    .card-head p{font-size:.82rem;opacity:.8;margin-top:3px}
    .card-body{padding:24px}

    /* Step indicator */
    .steps{display:flex;gap:0;margin-bottom:24px}
    .step{flex:1;text-align:center;padding:8px 4px;font-size:.75rem;color:#aaa;border-bottom:3px solid #e0e0e0;transition:.2s}
    .step.active{color:#1a5276;border-color:#1a5276;font-weight:600}
    .step.done{color:#28a745;border-color:#28a745}

    /* Drop zone */
    .dz{border:2px dashed #aab7c4;border-radius:8px;padding:40px 20px;text-align:center;cursor:pointer;transition:.2s;background:#f8f9fa}
    .dz:hover,.dz.over{border-color:#1a5276;background:#eaf0f8}
    .dz.loaded{border-color:#28a745;background:#f0fff4}
    .dz .icon{font-size:2.5rem;margin-bottom:8px}
    .dz p{color:#555;font-size:.9rem}
    .dz.loaded p{color:#28a745;font-weight:600}
    #fi{display:none}

    /* Sections */
    .sec-title{font-size:.75rem;font-weight:700;color:#1a5276;text-transform:uppercase;letter-spacing:.06em;
               margin:22px 0 10px;border-bottom:1px solid #e0e0e0;padding-bottom:6px;display:flex;align-items:center;gap:8px}
    .badge{font-size:.7rem;background:#e8f0fb;color:#1a5276;padding:2px 8px;border-radius:10px;font-weight:500;text-transform:none}
    .badge.warn{background:#fff3cd;color:#856404}
    .badge.ok{background:#d4edda;color:#155724}

    /* Fields */
    .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
    .full{grid-column:1/-1}
    .field label{display:block;font-size:.78rem;color:#444;margin-bottom:3px;font-weight:500}
    .field input,.field select{width:100%;padding:8px 11px;border:1px solid #ccc;border-radius:6px;font-size:.88rem;color:#222;transition:.15s}
    .field input:focus,.field select:focus{outline:none;border-color:#1a5276}
    .field input[readonly]{background:#f8f8f8;color:#666}
    .required::after{content:" *";color:#c0392b}

    /* Mapeo de parámetros sin asignar */
    .unmap-table{width:100%;border-collapse:collapse;font-size:.83rem}
    .unmap-table th{background:#f0f2f5;padding:7px 10px;text-align:left;color:#555;font-weight:600;border-bottom:2px solid #ddd}
    .unmap-table td{padding:6px 10px;border-bottom:1px solid #eee;vertical-align:middle}
    .unmap-table tr:last-child td{border-bottom:none}
    .unmap-table select{width:100%;padding:5px 8px;border:1px solid #ccc;border-radius:5px;font-size:.82rem}
    .unmap-table .skip-cb{width:16px;height:16px;cursor:pointer}
    .param-name{font-weight:500;color:#222}
    .param-cant{color:#888;font-size:.78rem}

    .alert{padding:12px 16px;border-radius:8px;font-size:.85rem;margin-bottom:12px}
    .alert.warn{background:#fff3cd;color:#856404;border:1px solid #ffc107}
    .alert.info{background:#d1ecf1;color:#0c5460;border:1px solid #bee5eb}

    /* Botón */
    .btn{display:block;width:100%;margin-top:20px;padding:12px;background:#1a5276;color:#fff;
         border:none;border-radius:8px;font-size:.95rem;font-weight:600;cursor:pointer;transition:.15s}
    .btn:hover{background:#154360}
    .btn:disabled{background:#aab7c4;cursor:not-allowed}
    .btn.secondary{background:#fff;color:#1a5276;border:2px solid #1a5276;margin-top:8px}
    .btn.secondary:hover{background:#eaf0f8}

    /* Status */
    .status{margin-top:14px;padding:12px 16px;border-radius:8px;font-size:.88rem;display:none}
    .status.loading{background:#eaf0f8;color:#1a5276;display:block}
    .status.success{background:#f0fff4;color:#1e8449;display:block}
    .status.error{background:#fdf2f8;color:#c0392b;display:block}
    .spinner{display:inline-block;width:13px;height:13px;border:2px solid currentColor;border-top-color:transparent;
             border-radius:50%;animation:spin .7s linear infinite;vertical-align:middle;margin-right:6px}
    @keyframes spin{to{transform:rotate(360deg)}}
    .dl-link{display:inline-block;margin-top:8px;color:#1a5276;font-weight:700;text-decoration:none;font-size:.95rem}
    .dl-link:hover{text-decoration:underline}

    /* Tabla editable de parámetros */
    .plan-table{width:100%;border-collapse:collapse;font-size:.83rem;margin-bottom:8px}
    .plan-table th{background:#1a5276;color:#fff;padding:7px 10px;text-align:left;font-weight:600;font-size:.78rem}
    .plan-table td{padding:4px 6px;border-bottom:1px solid #eee;vertical-align:middle}
    .plan-table tr:last-child td{border-bottom:none}
    .plan-table tr:hover td{background:#f8f9fa}
    .plan-table .item-num{text-align:center;color:#888;font-size:.78rem;font-weight:600}
    .plan-table select{width:100%;padding:4px 7px;border:1px solid #ccc;border-radius:4px;font-size:.82rem;background:#fff}
    .plan-table select:focus{outline:none;border-color:#1a5276}
    .plan-table .grupo-input{width:100%;padding:4px 7px;border:1px solid #ccc;border-radius:4px;font-size:.82rem;background:#fff}
    .plan-table .grupo-input:focus{outline:none;border-color:#1a5276}
    .plan-table input[type=number]{width:60px;padding:4px 6px;border:1px solid #ccc;border-radius:4px;font-size:.82rem;text-align:center}
    .plan-table input[type=number]:focus{outline:none;border-color:#1a5276}
    .btn-del{background:none;border:none;color:#c0392b;cursor:pointer;font-size:1rem;padding:2px 6px;border-radius:4px}
    .btn-del:hover{background:#fdf2f8}
    .btn-add-row{background:#fff;border:1px dashed #1a5276;color:#1a5276;padding:6px 14px;border-radius:6px;
                 font-size:.82rem;cursor:pointer;margin-top:4px;transition:.15s}
    .btn-add-row:hover{background:#eaf0f8}
  </style>
</head>
<body>
<div class="wrap">

  <div class="card">
    <div class="card-head">
      <h1>Generador de Plan de Muestreo</h1>
      <p>Cargue la cotización PDF para generar el plan automáticamente.</p>
    </div>
    <div class="card-body">

      <div class="steps">
        <div class="step active" id="s1">1. Cargar PDF</div>
        <div class="step" id="s2">2. Verificar datos</div>
        <div class="step" id="s3">3. Generar plan</div>
      </div>

      <!-- PASO 1: Drop zone -->
      <div id="paso1">
        <div class="dz" id="dz" onclick="document.getElementById('fi').click()">
          <div class="icon">📄</div>
          <p id="dz-txt">Haga clic o arrastre el PDF de la cotización aquí</p>
          <small style="color:#888;font-size:.78rem">Solo archivos .pdf</small>
        </div>
        <input type="file" id="fi" accept=".pdf">
      </div>

      <!-- PASO 2: Datos + parámetros sin mapeo -->
      <div id="paso2" style="display:none">

        <!-- Datos del cliente (del PDF, editables) -->
        <div class="sec-title">Datos de la cotización <span class="badge">Del PDF — editables</span></div>
        <div class="grid">
          <div class="field full"><label>Razón Social</label><input type="text" id="razon_social"></div>
          <div class="field full"><label>Dirección</label><input type="text" id="direccion"></div>
          <div class="field"><label>Contacto en campo</label><input type="text" id="contacto"></div>
          <div class="field"><label>Teléfono</label><input type="text" id="telefono"></div>
          <div class="field"><label>Email</label><input type="text" id="email"></div>
          <div class="field"><label>Matriz</label>
            <select id="matriz">
              <option value="">-- seleccionar --</option>
              {{MATRICES_OPTIONS}}
            </select>
          </div>
          <div class="field"><label>N° Cotización</label><input type="text" id="nro_cot" readonly></div>
          <div class="field"><label>Responsable cotización</label><input type="text" id="responsable" readonly></div>
        </div>

        <!-- Datos adicionales (manuales) -->
        <div class="sec-title">Datos adicionales <span style="color:#888;font-size:.72rem;font-weight:400">(no están en el PDF)</span></div>
        <div class="grid">
          <div class="field full"><label class="required">Lugar de muestreo</label><input type="text" id="lugar" placeholder="Ej: FUNDO SOCOSANI"></div>
          <div class="field"><label class="required">Preparado por</label><input type="text" id="prep_por" placeholder="Nombre del responsable"></div>
          <div class="field"><label>Fecha inicio muestreo</label><input type="date" id="fecha_ini"></div>
          <div class="field">
            <label>Muestreado por</label>
            <select id="muestreado_por">
              <option value="LABORATORIO">LABORATORIO</option>
              <option value="CLIENTE">CLIENTE</option>
            </select>
          </div>
          <div class="field full">
            <label>Observaciones <span style="color:#888;font-weight:400;font-size:.75rem">(puntos de muestreo, notas — editable)</span></label>
            <textarea id="observaciones" rows="3" style="width:100%;padding:8px 11px;border:1px solid #ccc;border-radius:6px;font-size:.88rem;resize:vertical;font-family:inherit"></textarea>
          </div>
        </div>

        <!-- Parámetros sin mapeo -->
        <div id="unmap-section" style="display:none">
          <div class="sec-title">
            Parámetros sin asignar
            <span class="badge warn" id="unmap-count"></span>
          </div>
          <div class="alert warn">
            Los siguientes ensayos del PDF no tienen grupo asignado en el plan.<br>
            Asígneles un grupo del Excel o márcelos como <strong>Ignorar</strong>.
            Las nuevas asignaciones se guardarán para futuras cotizaciones.
          </div>
          <table class="unmap-table">
            <thead><tr><th>Ensayo en la cotización</th><th>Muestras</th><th>Grupo en el plan</th><th>Ignorar</th></tr></thead>
            <tbody id="unmap-tbody"></tbody>
          </table>
        </div>

        <!-- datalist compartido para autocompletar grupos -->
        <datalist id="grupos-list">{{GRUPOS_DATALIST}}</datalist>

        <!-- Tabla editable de parámetros del plan -->
        <div id="plan-section" style="display:none">
          <div class="sec-title">
            Parámetros del plan
            <span class="badge ok" id="plan-count"></span>
            <span style="margin-left:auto;font-size:.72rem;color:#888;font-weight:400">Editable — seleccione o escriba libremente</span>
          </div>
          <table class="plan-table" id="plan-table">
            <thead>
              <tr>
                <th style="width:36px">#</th>
                <th>Grupo (Parámetros)</th>
                <th style="width:80px;text-align:center">Puntos</th>
                <th style="width:36px"></th>
              </tr>
            </thead>
            <tbody id="plan-tbody"></tbody>
          </table>
          <button class="btn-add-row" onclick="agregarFila()">+ Agregar fila</button>
        </div>

        <!-- Sistema de Control de Calidad -->
        <div id="qc-section" style="display:none">
          <div class="sec-title">
            Sistema de Control de Calidad
            <span style="margin-left:auto;font-size:.72rem;color:#888;font-weight:400">Escoja el parámetro de cada blanco/duplicado</span>
          </div>
          <table class="plan-table" id="qc-table">
            <thead>
              <tr>
                <th style="width:180px">Control de calidad</th>
                <th>Parámetros</th>
                <th style="width:80px">Envase</th>
                <th style="width:90px">Volumen</th>
                <th style="width:150px">Preservación</th>
              </tr>
            </thead>
            <tbody id="qc-tbody"></tbody>
          </table>
        </div>

        <button class="btn" id="btn-gen" onclick="generarPlan()">Generar Plan de Muestreo</button>
        <button class="btn secondary" onclick="recargar()">← Cargar otra cotización</button>
      </div>

      <div class="status" id="status"></div>
    </div>
  </div>
</div>

<script>
const GRUPOS = {{GRUPOS_JSON}};
const PARAMS_INFO = {{PARAMS_INFO_JSON}};
const QC_RENGLONES = [
  "Blanco de Campo (BK-C)",
  "Blanco Viajero (BK-V)",
  "Duplicado (Dup)",
  "Duplicado (Dup-MB)",
  "Blanco Viajero MB (BK-VM)"
];
let pdfFile = null;
let todosParametros = [];

// ── Drop zone ──────────────────────────────────────────────────────────────
const dz = document.getElementById('dz');
const fi = document.getElementById('fi');
dz.addEventListener('dragover', e=>{e.preventDefault();dz.classList.add('over')});
dz.addEventListener('dragleave', ()=>dz.classList.remove('over'));
dz.addEventListener('drop', e=>{e.preventDefault();dz.classList.remove('over');const f=e.dataTransfer.files[0];if(f&&f.name.endsWith('.pdf'))cargarPDF(f)});
fi.addEventListener('change', ()=>{if(fi.files[0])cargarPDF(fi.files[0])});

function cargarPDF(file) {
  pdfFile = file;
  dz.classList.add('loaded');
  document.getElementById('dz-txt').textContent = file.name;
  setStatus('loading', 'Leyendo cotización...');

  const fd = new FormData();
  fd.append('pdf', file);
  fetch('/leer_pdf', {method:'POST', body:fd})
    .then(r=>r.json())
    .then(data=>{
      if(data.error){setStatus('error','Error: '+data.error);return}
      rellenarFormulario(data);
      clearStatus();
      paso(2);
    })
    .catch(e=>setStatus('error','Error al leer PDF: '+e.message));
}

function rellenarFormulario(data) {
  document.getElementById('razon_social').value   = data.razon_social||'';
  document.getElementById('direccion').value      = data.direccion||'';
  document.getElementById('contacto').value       = data.contacto||'';
  document.getElementById('telefono').value       = data.telefono_contacto||'';
  document.getElementById('email').value          = data.email_contacto||'';
  document.getElementById('nro_cot').value        = data.nro_cotizacion||'';
  document.getElementById('responsable').value    = data.responsable_cotizacion||'';
  document.getElementById('observaciones').value  = data.informacion_adicional||'';

  // Pre-seleccionar la matriz: buscar coincidencia exacta o parcial
  const matrizPDF = (data.matriz || '').toUpperCase();
  const selMatriz = document.getElementById('matriz');
  let mejorOpcion = '';
  [...selMatriz.options].forEach(opt => {
    const optVal = opt.value.toUpperCase();
    // Coincidencia exacta
    if(optVal === matrizPDF) { mejorOpcion = opt.value; }
    // Coincidencia parcial: el valor del PDF está contenido en la opción o viceversa
    else if(!mejorOpcion && (optVal.includes(matrizPDF) || matrizPDF.includes(optVal.replace(/^\d+\.\s*/,'').split('/')[0].trim()))) {
      mejorOpcion = opt.value;
    }
  });
  selMatriz.value = mejorOpcion;

  todosParametros = data.parametros || [];

  // Obtener mapeo + grupos resueltos
  fetch('/mapeo_check', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({parametros: todosParametros.map(p=>p.nombre),
                          n_muestras: data.n_muestras || ''})
  })
  .then(r=>r.json())
  .then(res=>{
    mostrarParametros(res.sin_mapeo);
    poblarTablaPlan(res.grupos_resueltos || []);
  });
}

function mostrarParametros(sinMapeo) {
  const sec   = document.getElementById('unmap-section');
  const tbody = document.getElementById('unmap-tbody');
  tbody.innerHTML = '';
  if(!sinMapeo || sinMapeo.length === 0){ sec.style.display='none'; return; }

  sec.style.display='block';
  document.getElementById('unmap-count').textContent = sinMapeo.length + ' sin asignar';
  const opts = GRUPOS.map(g=>`<option value="${g}">${g}</option>`).join('');
  sinMapeo.forEach(item=>{
    const tr = document.createElement('tr');
    tr.innerHTML = `
      <td><span class="param-name">${item.nombre}</span></td>
      <td><span class="param-cant">${item.cantidad||'—'}</span></td>
      <td><select data-nombre="${item.nombre}">
        <option value="">-- seleccionar grupo --</option>
        <option value="SKIP">[ Ignorar este ensayo ]</option>
        ${opts}
      </select></td>
      <td style="text-align:center">
        <input type="checkbox" class="skip-cb" onchange="toggleSkip(this,'${item.nombre.replace(/'/g,"\\'")}')">
      </td>`;
    tbody.appendChild(tr);
  });
}

function poblarTablaPlan(grupos) {
  const sec   = document.getElementById('plan-section');
  const tbody = document.getElementById('plan-tbody');
  tbody.innerHTML = '';
  sec.style.display = grupos.length ? 'block' : 'none';
  document.getElementById('plan-count').textContent = grupos.length + ' filas';

  grupos.forEach((item, i) => agregarFilaConDatos(item.grupo, item.puntos));
  renumerarFilas();

  poblarTablaQC();
}

function poblarTablaQC() {
  const sec   = document.getElementById('qc-section');
  const tbody = document.getElementById('qc-tbody');
  tbody.innerHTML = '';
  sec.style.display = 'block';

  QC_RENGLONES.forEach((etiqueta, i) => {
    const tr = document.createElement('tr');
    tr.innerHTML = `
      <td style="font-weight:600;color:#333;font-size:.8rem">${etiqueta}</td>
      <td>
        <input type="text" class="grupo-input qc-grupo" list="grupos-list"
               data-row="${i}" oninput="actualizarQC(${i})"
               placeholder="Escoja o escriba un parámetro...">
      </td>
      <td class="qc-envase"  style="text-align:center;color:#555;font-size:.8rem">—</td>
      <td class="qc-volumen" style="text-align:center;color:#555;font-size:.8rem">—</td>
      <td class="qc-preserv" style="text-align:center;color:#555;font-size:.8rem">—</td>`;
    tbody.appendChild(tr);
  });
}

function actualizarQC(i) {
  const tr  = document.querySelectorAll('#qc-tbody tr')[i];
  const val = tr.querySelector('.qc-grupo').value.trim();
  const info = PARAMS_INFO[val];
  tr.querySelector('.qc-envase').textContent  = info ? info.envase       : '—';
  tr.querySelector('.qc-volumen').textContent = info ? info.volumen      : '—';
  tr.querySelector('.qc-preserv').textContent = info ? info.preservacion : '—';
}

function agregarFilaConDatos(grupo='', puntos='') {
  const tbody = document.getElementById('plan-tbody');
  const tr    = document.createElement('tr');
  tr.innerHTML = `
    <td class="item-num">—</td>
    <td>
      <input type="text" class="grupo-input" list="grupos-list"
             value="${escH(String(grupo))}"
             placeholder="Seleccione o escriba un grupo...">
    </td>
    <td style="text-align:center">
      <input type="number" class="puntos-inp" min="1" max="99" value="${escH(String(puntos))}" style="width:60px">
    </td>
    <td><button class="btn-del" onclick="eliminarFila(this)" title="Eliminar fila">✕</button></td>`;
  tbody.appendChild(tr);
}

function agregarFila() {
  agregarFilaConDatos('', document.querySelector('.puntos-inp')?.value || 1);
  renumerarFilas();
}

function eliminarFila(btn) {
  btn.closest('tr').remove();
  renumerarFilas();
}

function renumerarFilas() {
  document.querySelectorAll('#plan-tbody tr').forEach((tr, i)=>{
    tr.querySelector('.item-num').textContent = i + 1;
  });
  const n = document.querySelectorAll('#plan-tbody tr').length;
  const el = document.getElementById('plan-count');
  if(el) el.textContent = n + ' filas';
}

function escH(s){ return String(s).replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/</g,'&lt;'); }

function toggleSkip(cb, nombre) {
  const sel = document.querySelector(`select[data-nombre="${nombre}"]`);
  if(sel) sel.value = cb.checked ? 'SKIP' : '';
}

// ── Generar plan ───────────────────────────────────────────────────────────
function generarPlan() {
  const lugar   = document.getElementById('lugar').value.trim();
  const prepPor = document.getElementById('prep_por').value.trim();
  if(!lugar||!prepPor){setStatus('error','Complete los campos obligatorios (*).');return}

  const fechaVal = document.getElementById('fecha_ini').value;
  let fechaStr = '';
  if(fechaVal){const[y,m,d]=fechaVal.split('-');fechaStr=`${d}/${m}/${y}`}

  // Recoger asignaciones de parámetros sin mapeo
  const nuevosMapeos = {};
  document.querySelectorAll('#unmap-tbody select').forEach(sel=>{
    if(sel.value) nuevosMapeos[sel.dataset.nombre] = sel.value;
  });

  // Verificar que todos tienen asignación
  const sinAsignar = [...document.querySelectorAll('#unmap-tbody select')]
    .filter(s=>!s.value).map(s=>s.dataset.nombre);
  if(sinAsignar.length>0){
    setStatus('error',`Asigne un grupo (o marque Ignorar) para: ${sinAsignar.join(', ')}`);
    return;
  }

  // Recoger tabla editable de parámetros
  const gruposEditados = [];
  document.querySelectorAll('#plan-tbody tr').forEach(tr=>{
    const grupo  = tr.querySelector('.grupo-input').value.trim();
    const puntos = parseInt(tr.querySelector('.puntos-inp').value) || 1;
    if(grupo) gruposEditados.push({grupo, puntos});
  });
  if(gruposEditados.length === 0){
    setStatus('error','La tabla de parámetros está vacía. Agregue al menos un grupo.');
    return;
  }

  document.getElementById('btn-gen').disabled = true;
  setStatus('loading','<span class="spinner"></span> Generando plan de muestreo...');

  const fd = new FormData();
  fd.append('pdf', pdfFile);
  fd.append('preparado_por',  prepPor);
  fd.append('lugar_muestreo', lugar);
  fd.append('fecha_inicio',   fechaStr);
  fd.append('razon_social',   document.getElementById('razon_social').value);
  fd.append('direccion',      document.getElementById('direccion').value);
  fd.append('contacto',       document.getElementById('contacto').value);
  fd.append('telefono',       document.getElementById('telefono').value);
  fd.append('email',          document.getElementById('email').value);
  fd.append('observaciones',   document.getElementById('observaciones').value);
  fd.append('muestreado_por', document.getElementById('muestreado_por').value);
  fd.append('matriz',         document.getElementById('matriz').value);

  // Sistema de control de calidad (5 renglones fijos)
  const controlCalidad = [];
  document.querySelectorAll('#qc-tbody tr').forEach(tr=>{
    controlCalidad.push(tr.querySelector('.qc-grupo').value.trim());
  });
  fd.append('control_calidad', JSON.stringify(controlCalidad));
  fd.append('nuevos_mapeos',  JSON.stringify(nuevosMapeos));
  fd.append('grupos_editados', JSON.stringify(gruposEditados));

  fetch('/generar',{method:'POST',body:fd})
    .then(r=>r.json())
    .then(data=>{
      document.getElementById('btn-gen').disabled=false;
      if(data.error){setStatus('error','Error: '+data.error);return}
      paso(3);
      const guardados = data.nuevos_mapeos_guardados||0;
      setStatus('success',
        `Plan generado correctamente.${guardados>0?' Se guardaron '+guardados+' nuevo(s) mapeo(s).':''}<br>
         <a class="dl-link" href="/descargar/${data.filename}">⬇ Descargar ${data.filename}</a>`);
    })
    .catch(e=>{document.getElementById('btn-gen').disabled=false;setStatus('error','Error: '+e.message)});
}

// ── Helpers ────────────────────────────────────────────────────────────────
function paso(n){
  [1,2,3].forEach(i=>{
    const el=document.getElementById('s'+i);
    el.className='step'+(i<n?' done':i===n?' active':'');
  });
  document.getElementById('paso1').style.display=n===1?'block':'none';
  document.getElementById('paso2').style.display=n>=2?'block':'none';
}
function setStatus(t,h){const e=document.getElementById('status');e.className='status '+t;e.innerHTML=h}
function clearStatus(){const e=document.getElementById('status');e.className='status';e.innerHTML=''}
function recargar(){location.reload()}
</script>
</body>
</html>
"""


# ── Rutas Flask ───────────────────────────────────────────────────────────────

@app.route("/")
def index():
    grupos_list = grupos_disponibles()
    grupos      = json.dumps(grupos_list, ensure_ascii=False)
    info_json   = json.dumps(params_info(), ensure_ascii=False)
    matrices    = matrices_disponibles()
    mat_opts    = "\n".join(f'<option value="{m}">{m}</option>' for m in matrices)
    dl_opts     = "\n".join(f'<option value="{g}">' for g in grupos_list)
    html = HTML.replace("{{GRUPOS_JSON}}", grupos)
    html = html.replace("{{PARAMS_INFO_JSON}}", info_json)
    html = html.replace("{{MATRICES_OPTIONS}}", mat_opts)
    html = html.replace("{{GRUPOS_DATALIST}}", dl_opts)
    return render_template_string(html)


@app.route("/leer_pdf", methods=["POST"])
def leer_pdf():
    if "pdf" not in request.files:
        return jsonify({"error": "No se recibió ningún archivo."})
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        request.files["pdf"].save(tmp.name)
        tmp_path = tmp.name
    try:
        return jsonify(extraer_cotizacion(tmp_path))
    except Exception as e:
        return jsonify({"error": str(e)})
    finally:
        os.unlink(tmp_path)


@app.route("/mapeo_check", methods=["POST"])
def mapeo_check():
    """Devuelve parámetros sin mapeo y grupos resueltos con sus puntos."""
    from generar_plan import resolver_grupos, extraer_n_puntos

    payload    = request.json or {}
    parametros = payload.get("parametros", [])
    n_muestras = payload.get("n_muestras", "")

    with open(MAPEO_PATH, encoding="utf-8") as f:
        mapeo = json.load(f)

    sin_mapeo = [{"nombre": n, "cantidad": "—"}
                 for n in parametros if n not in mapeo]

    # Resolver grupos usando la lógica de familias
    grupos_plan = resolver_grupos(parametros, mapeo)

    # Determinar puntos por defecto
    try:
        n_puntos = int(n_muestras) if n_muestras else 1
    except (ValueError, TypeError):
        n_puntos = 1

    grupos_resueltos = [{"grupo": g, "puntos": n_puntos} for g in grupos_plan]

    return jsonify({"sin_mapeo": sin_mapeo, "grupos_resueltos": grupos_resueltos})


@app.route("/generar", methods=["POST"])
def generar():
    if "pdf" not in request.files:
        return jsonify({"error": "No se recibió ningún archivo."})

    nuevos_mapeos = json.loads(request.form.get("nuevos_mapeos", "{}"))
    overrides = {
        "razon_social": request.form.get("razon_social"),
        "direccion":    request.form.get("direccion"),
        "contacto":     request.form.get("contacto"),
        "telefono":     request.form.get("telefono"),
        "email":        request.form.get("email"),
        "matriz":       request.form.get("matriz"),
    }

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        request.files["pdf"].save(tmp.name)
        tmp_path = tmp.name

    try:
        datos = extraer_cotizacion(tmp_path)
        nro = nro_base(datos["nro_cotizacion"])
        cliente = datos["razon_social"].replace(" ", "_")[:20]
        filename = f"PM_{nro}_{cliente}.xlsm"
        ruta_salida = OUTPUT_DIR / filename

        # Guardar nuevos mapeos en el JSON antes de generar
        guardados = 0
        if nuevos_mapeos:
            with open(MAPEO_PATH, encoding="utf-8") as f:
                mapeo = json.load(f)
            for nombre, grupo in nuevos_mapeos.items():
                if nombre not in mapeo or mapeo[nombre] != grupo:
                    mapeo[nombre] = grupo
                    guardados += 1
            with open(MAPEO_PATH, "w", encoding="utf-8") as f:
                json.dump(mapeo, f, ensure_ascii=False, indent=2)

        muestreado_por = request.form.get("muestreado_por", "LABORATORIO")
        grupos_editados_raw = request.form.get("grupos_editados", "[]")
        grupos_override = json.loads(grupos_editados_raw) or None
        control_calidad = json.loads(request.form.get("control_calidad", "[]")) or None

        salida = generar_plan(
            ruta_pdf=tmp_path,
            preparado_por=request.form.get("preparado_por", ""),
            lugar_muestreo=request.form.get("lugar_muestreo", ""),
            fecha_inicio=request.form.get("fecha_inicio", ""),
            observaciones=request.form.get("observaciones", ""),
            muestreado_por=muestreado_por,
            ruta_salida=str(ruta_salida),
            overrides=overrides,
            grupos_override=grupos_override,
            control_calidad=control_calidad,
        )
        return jsonify({"filename": salida.name, "nuevos_mapeos_guardados": guardados})
    except Exception as e:
        import traceback
        return jsonify({"error": traceback.format_exc()})
    finally:
        os.unlink(tmp_path)


@app.route("/descargar/<filename>")
def descargar(filename):
    ruta = OUTPUT_DIR / filename
    if not ruta.exists():
        return "Archivo no encontrado", 404
    return send_file(str(ruta), as_attachment=True, download_name=filename,
                     mimetype="application/vnd.ms-excel.sheet.macroenabled.12")


if __name__ == "__main__":
    print("\n  Plan de Muestreo — Generador")
    print("  Abre tu navegador en: http://localhost:5003\n")
    app.run(debug=False, port=5003)
