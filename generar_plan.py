"""
Paso 3: Genera el Plan de Muestreo Excel a partir de una cotización PDF.
Uso: python generar_plan.py <ruta_pdf> [opciones]
"""

import json
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

from extraer_cotizacion import extraer_cotizacion

# ─── Orden canónico de los grupos en el plan ──────────────────────────────────
ORDEN_GRUPOS = [
    # Microbiológico
    "COLIF.FECALES*/TOTALES* - E.COLI* - BACT.HETEROTROFAS* - VIRUS",
    "COLIFORMES FECALES*-TOTALES*",
    "COLIFORMES TOTALES* - BACT.HETEROTROFAS*",
    "ORGANISMOS DE VIDA LIBRE (O.V.L)",
    "HUEVOS Y LARVAS HELMINTOS",
    "FORMAS PARASITARIAS",
    "PSEUDOMONAS AERUGINOSA",
    "Recuento de microorganismos aerobios mesófilos",
    "Recuento de mohos y levaduras",
    "ENUMERACIÓN DE BACTERIAS ANAEROBIAS SULFITO REDUCTORES",
    "ESTREPTOCOCOS FECALES",
    "ENUMERACION DE ENTEROCOCOS",
    # Físico-químico básico
    "COLOR - OLOR - SABOR",
    "PH* - T° - CONDUCTIVIDAD - TURBIDEZ - CLORO LIBRE - O.D (Campo)",
    "SOLIDOS TOTALES DISUELTOS (TDS)*",
    "SOLIDOS SUSPENDIDOS TOTALES (TSS)*",
    "SOLIDOS TOTALES (ST)*",
    "SOLIDOS SEDIMENTABLES (SS)*",
    # Demandas
    "DEMANDA BIOQUIMICA DE OXIGENO (DBO5)*",
    "DEMANDA QUIMICA DE OXIGENO (DQO)*",
    # Iones y nutrientes
    "CLORUROS - SULFATOS* - FLUOR",
    "CLORITO - CLORATO - NITRATOS - NITRITOS - BROMATOS",
    "NITRATOS - NITRITOS",
    "NITROGENO AMONIACAL*",
    "NITROGENO ORGANICO Kjeldahl",
    "NITROGENO TOTAL",
    "FOSFORO TOTAL",
    "FOSFATOS",
    "FLUORUROS ",
    "ALCALINIDAD*",
    "SULFUROS*",
    "AMONIACO",
    "AMONIO (NITROGENO AMONIACAL)",
    "SILICE, SILICATO",
    "POTASIO - SODIO",
    "CALCIO - MAGNESIO - POTASIO - SODIO",
    # Metales
    "METALES TOTALES AA* e ICP-OES - DUREZA TOTAL",
    "METALES TOTALES AA* e ICP-OES",
    "DUREZA TOTAL",
    "DUREZA CALCICA",
    "CIANURO TOTAL",
    "CN- TOTAL* (CIANURO TOTAL)",
    "CROMO HEXAVALENTE*",
    "MERCURIO",
    # Orgánicos
    "ACEITES Y GRASAS*",
    "Hidrocarburos totales (TPH; F2; F3)",
    "PCBs",
    "MCPA",
    "MICROSISTINA - LR",
    "Detergentes SAAM",
    "ORGANICOS - INORGANICOS",
    # Otros
    "TEMPERATURA",
    "TEMPERATURA* (Campo)",
]

import tempfile as _tempfile

PLANTILLA      = Path(__file__).resolve().parent / "COT2026-0103 - SOCOSANI S A.xlsm"
MAPEO_PATH     = Path(__file__).resolve().parent / "mapeo_parametros.json"
FAMILIAS_PATH  = Path(__file__).resolve().parent / "familias_parametros.json"

def _dir_escribible(path: Path) -> Path:
    """Devuelve el path si es escribible, o /tmp como fallback."""
    try:
        path.mkdir(exist_ok=True)
        t = path / ".write_test"
        t.touch(); t.unlink()
        return path
    except OSError:
        fb = Path(_tempfile.gettempdir()) / "planes_generados"
        fb.mkdir(exist_ok=True)
        return fb


def resolver_grupos(nombres_pdf: list[str], mapeo: dict) -> list[str]:
    """
    Determina los grupos del plan de muestreo a partir de los parámetros del PDF.

    Para parámetros que pertenecen a una 'familia' (combinables), elige el grupo
    exacto según qué miembros de la familia están presentes en la cotización.
    Para el resto usa el mapeo directo.

    Returns:
        Lista ordenada de nombres de grupos del Excel (únicos, en orden canónico).
    """
    with open(FAMILIAS_PATH, encoding="utf-8") as f:
        familias_def = json.load(f)["familias"]

    grupos_resultado: set[str] = set()
    parametros_procesados: set[str] = set()

    # 1. Procesar familias con COBERTURA GREEDY POR SUBCONJUNTO.
    #    Regla de oro: un grupo combinado solo se usa si TODOS sus miembros
    #    están presentes en la cotización (kset ⊆ tokens_presentes).
    #    Se elige el combo válido más grande, se descuentan sus tokens y se
    #    repite hasta cubrir todo. Cada token debe tener un grupo individual.
    for familia in familias_def:
        indiv_a_token: dict = familia["individual_a_token"]
        tokens_a_grupo: dict = familia["tokens_a_grupo"]

        tokens_presentes: set[str] = set()
        miembros_en_cot: set[str] = set()
        for param in nombres_pdf:
            if param in indiv_a_token:
                tokens_presentes.add(indiv_a_token[param])
                miembros_en_cot.add(param)

        if not tokens_presentes:
            continue

        # Precalcular combos como (set_de_tokens, grupo), ordenados de mayor a menor
        combos = [(frozenset(k.split("|")), g) for k, g in tokens_a_grupo.items()]
        combos.sort(key=lambda x: len(x[0]), reverse=True)

        restante = set(tokens_presentes)
        cubiertos: set[str] = set()
        while restante:
            elegido = None
            for kset, grupo in combos:
                # SOLO válido si todos los tokens del combo están presentes
                if kset.issubset(restante):
                    elegido = (kset, grupo)
                    break
            if elegido is None:
                break  # tokens sin combo → caerán al mapeo directo
            kset, grupo = elegido
            grupos_resultado.add(grupo)
            cubiertos |= kset
            restante -= kset

        # Marcar como procesados SOLO los params cuyo token quedó cubierto
        for param in miembros_en_cot:
            if indiv_a_token[param] in cubiertos:
                parametros_procesados.add(param)

    # 2. Procesar los parámetros restantes con mapeo directo
    for param in nombres_pdf:
        if param in parametros_procesados:
            continue
        grupo = mapeo.get(param)
        if grupo and grupo != "SKIP":
            grupos_resultado.add(grupo)

    # 3. Ordenar según el orden canónico
    ordenados = [g for g in ORDEN_GRUPOS if g in grupos_resultado]
    for g in grupos_resultado:
        if g not in ordenados:
            ordenados.append(g)

    return ordenados


def nro_base(nro: str) -> str:
    """
    Devuelve el número de cotización sin el sufijo de revisión, robusto ante
    correlativos de CUALQUIER cantidad de dígitos:
       '2026-0103-3' -> '2026-0103'   (quita revisión)
       '2026-0410'   -> '2026-0410'   (sin cambio)
       '2025-11656'  -> '2025-11656'  (5 dígitos, sin cambio)
       '2025-11656-2'-> '2025-11656'  (quita revisión)
    Regla: si hay 3+ segmentos separados por '-', el último es la revisión.
    """
    partes = (nro or "").split("-")
    if len(partes) >= 3:
        return "-".join(partes[:2])
    return nro or ""


def extraer_responsable(nombre_completo: str) -> str:
    """'ANGELA ALEJANDRA DELGADO BEDREGAL' → 'ALEJANDRA' (segundo nombre)."""
    palabras = nombre_completo.strip().split()
    return palabras[1] if len(palabras) > 1 else palabras[0] if palabras else ""


def extraer_n_puntos(info_adicional: str) -> int:
    """Extrae el número de vertientes/puntos del texto de información adicional."""
    m = re.search(r"(\d+)\s+VERTIENTES", info_adicional, re.IGNORECASE)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\s+(PUNTOS|MUESTRAS|ESTACIONES)", info_adicional, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return 1


def generar_plan(
    ruta_pdf: str,
    preparado_por: str,
    lugar_muestreo: str,
    fecha_inicio: str,
    observaciones: str = "",
    muestreado_por: str = "LABORATORIO",
    ruta_salida: str | None = None,
    overrides: dict | None = None,
    grupos_override: list | None = None,   # [{"grupo": str, "puntos": int}]
    control_calidad: list | None = None,   # [bk_c, bk_v, dup, dup_mb, bk_vm]
) -> bytes:
    """
    Genera el Excel del plan de muestreo a partir del PDF de cotización.

    Args:
        ruta_pdf:       Ruta al PDF de la cotización.
        preparado_por:  Nombre del responsable interno del plan.
        lugar_muestreo: Ubicación del muestreo (no está en el PDF).
        fecha_inicio:   Fecha de inicio del muestreo (dd/mm/yyyy).
        ruta_salida:    Ruta del archivo Excel de salida (opcional).

    Returns:
        Path al archivo Excel generado.
    """
    # 1. Extraer datos del PDF
    datos = extraer_cotizacion(ruta_pdf)

    # Aplicar sobreescrituras del usuario (campos editados en la interfaz)
    if overrides:
        for campo, valor in overrides.items():
            if valor is not None:
                datos[campo] = valor

    # 2. Cargar mapeo
    with open(MAPEO_PATH, encoding="utf-8") as f:
        mapeo = json.load(f)

    # 3. Determinar grupos
    if grupos_override is not None:
        # El usuario editó la tabla en la interfaz — usar esos valores directamente
        grupos_plan   = [item["grupo"]  for item in grupos_override]
        puntos_override = {item["grupo"]: int(item.get("puntos", 1)) for item in grupos_override}
    else:
        nombres_pdf = [p["nombre"] for p in datos["parametros"]]
        grupos_plan = resolver_grupos(nombres_pdf, mapeo)
        puntos_override = None

    n_puntos = extraer_n_puntos(datos["informacion_adicional"])
    # Si info_adicional está vacía o no tiene vertientes, usar n_muestras del PDF
    if n_puntos <= 1 and datos.get("n_muestras"):
        try:
            n_puntos = int(datos["n_muestras"])
        except (ValueError, TypeError):
            pass

    # 4. Cargar la plantilla EN MEMORIA (sin copiar a disco — compatible con
    #    entornos de solo lectura como Vercel/serverless).
    wb = openpyxl.load_workbook(PLANTILLA, keep_vba=True)
    ws = wb["PLAN DE MUESTREO-AGUA"]

    # 5. Número de cotización (sin sufijo de revisión, con "/" como en la plantilla)
    # "2026-0103-3" → "2026/0103"  |  "2025-11656" → "2025/11656"
    ws["Y2"] = nro_base(datos["nro_cotizacion"]).replace("-", "/")

    # 6. Cabecera del plan
    ws["C3"] = preparado_por
    ws["P3"] = datetime.today()

    # 7. Datos del cliente
    ws["F6"] = datos.get("razon_social", "")
    ws["F7"] = datos.get("direccion", "")
    ws["F8"] = lugar_muestreo
    ws["F9"] = datos.get("contacto", "")
    ws["X9"] = "RESPONSABLE"
    ws["F10"] = datos.get("telefono", "") or datos.get("telefono_contacto", "")
    ws["X10"] = datos.get("email", "") or datos.get("email_contacto", "")
    ws["F11"] = fecha_inicio
    ws["X11"] = extraer_responsable(datos["responsable_cotizacion"])
    ws["X12"] = muestreado_por.upper() if muestreado_por else "LABORATORIO"

    # 8. Tipo de matriz
    ws["C15"] = datos.get("matriz", "")

    # 9. Limpiar filas de parámetros (17-47) y rellenar
    for row in range(17, 48):
        ws.cell(row=row, column=2).value = None   # B = nombre grupo
        ws.cell(row=row, column=16).value = None  # P = puntos

    for i, grupo in enumerate(grupos_plan):
        if i >= 31:
            break
        row = 17 + i
        ws.cell(row=row, column=2).value = grupo
        pts = puntos_override.get(grupo, n_puntos) if puntos_override else n_puntos
        ws.cell(row=row, column=16).value = pts

    # 10. Observaciones
    obs = observaciones.strip() if observaciones.strip() else datos.get("informacion_adicional", "").rstrip(".")
    ws["A49"] = obs

    # 11. Sistema de Control de Calidad (filas 77-81, columna F = col 6).
    #     Se escriben los parámetros que el usuario escogió en la interfaz.
    #     Las fórmulas VLOOKUP del Excel completan envase/volumen/preservación.
    if control_calidad:
        filas_qc = [77, 78, 79, 80, 81]  # BK-C, BK-V, Dup, Dup-MB, BK-VM
        for fila, valor in zip(filas_qc, control_calidad):
            ws.cell(row=fila, column=6).value = (valor or "").strip()
        # Limpiar renglones no provistos
        for fila in filas_qc[len(control_calidad):]:
            ws.cell(row=fila, column=6).value = ""

    # 12. Guardar EN MEMORIA y devolver los bytes del .xlsm.
    #     (No se escribe en disco salvo que se pida ruta_salida explícita.)
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    data = buffer.getvalue()

    # Uso local / CLI: si se pidió guardar en disco, intentarlo sin romper si falla.
    if ruta_salida is not None:
        try:
            destino = Path(ruta_salida)
            destino = _dir_escribible(destino.parent) / destino.name
            destino.write_bytes(data)
        except OSError:
            pass

    return data


# ─── Ejecución directa ────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python generar_plan.py <ruta_pdf>")
        print("     Se pedirán los campos adicionales interactivamente.")
        sys.exit(1)

    ruta_pdf = sys.argv[1]

    print("\n─── Datos adicionales (no están en el PDF) ───")
    preparado_por  = input("  Preparado por       : ").strip()
    lugar_muestreo = input("  Lugar de muestreo   : ").strip()
    fecha_inicio   = input("  Fecha inicio muestreo (dd/mm/yyyy): ").strip()

    nombre = f"PM_generado.xlsm"
    generar_plan(ruta_pdf, preparado_por, lugar_muestreo, fecha_inicio, ruta_salida=nombre)
    print(f"\nPlan generado: {nombre}")
